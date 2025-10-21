"""Jira Worklog CLI (Python)

Implements similar behavior to the provided Java tool. Filters worklogs by EMAIL env and by period.
"""
from __future__ import annotations
import os
import sys
import csv
import base64
import requests
from urllib.parse import quote_plus
from datetime import datetime, date, time, timedelta
from dateutil import parser
import pytz
from typing import Optional, Tuple, List, Dict

# load .env if present
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    # dotenv is optional at import-time; installation ensures it's present
    pass

DEFAULT_BASE = "https://subex.atlassian.net"


class IssueTotal:
    def __init__(self, key: str, summary: str, project: str, total_seconds: int):
        self.key = key
        self.summary = summary
        self.project = project
        self.total_seconds = total_seconds

    def hours(self):
        return self.total_seconds / 3600.0

    def days(self):
        return self.hours() / 8.0


def _flatten_text(node) -> str:
    """Recursively extract text from nested Jira/ADF comment structures.

    Handles strings, dicts with 'text' fields, lists of such nodes, and nested 'content' lists.
    Returns a plain string.
    """
    if node is None:
        return ''
    if isinstance(node, str):
        return node
    texts: List[str] = []
    if isinstance(node, dict):
        # direct text field
        if 'text' in node and isinstance(node['text'], str):
            texts.append(node['text'])
        # common ADF shape: { 'type': 'paragraph', 'content': [ ... ] }
        for key in ('content', 'children', 'body'):
            if key in node and isinstance(node[key], (list, dict)):
                texts.append(_flatten_text(node[key]))
        # legacy shapes
        for v in node.values():
            if isinstance(v, (list, dict, str)):
                # avoid duplicating text fields
                if v is node.get('text'):
                    continue
                texts.append(_flatten_text(v))
    elif isinstance(node, list):
        for item in node:
            texts.append(_flatten_text(item))
    else:
        # fallback to string conversion
        try:
            return str(node)
        except Exception:
            return ''
    # join and normalize whitespace
    return ' '.join(t.strip() for t in texts if t and isinstance(t, str)).strip()


def compute_range(period: str, tz: Optional[pytz.timezone] = None) -> Optional[Tuple[datetime, datetime]]:
    if tz is None:
        tz = pytz.timezone(os.environ.get('TZ', 'UTC'))
    today = datetime.now(tz).date()
    if period == 'this':
        naive_start = datetime.combine(today.replace(day=1), time.min)
        last_day = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        naive_end = datetime.combine(last_day, time.max)
        start = tz.localize(naive_start)
        end = tz.localize(naive_end)
        return (start, end)
    if period == 'last':
        first_of_this = today.replace(day=1)
        last_month_last_day = first_of_this - timedelta(days=1)
        naive_start = datetime.combine(last_month_last_day.replace(day=1), time.min)
        naive_end = datetime.combine(last_month_last_day, time.max)
        start = tz.localize(naive_start)
        end = tz.localize(naive_end)
        return (start, end)
    return None


def escape(s: Optional[str]) -> str:
    if s is None:
        return ''
    s = s.replace('"', '""')
    if ',' in s or '"' in s or '\n' in s:
        return f'"{s}"'
    return s


def get_env_or_prompt(name: str, prompt_text: str, hide: bool = False) -> str:
    v = os.environ.get(name)
    if v:
        return v
    try:
        if hide:
            import getpass

            return getpass.getpass(prompt_text)
        else:
            return input(prompt_text)
    except KeyboardInterrupt:
        print()
        sys.exit(1)


def search_issues(base: str, email: str, api_token: str, jql: str) -> List[dict]:
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic ' + base64.b64encode(f"{email}:{api_token}".encode('utf-8')).decode('utf-8')
    }
    print("Headers: ", headers)
    all_issues = []
    start_at = 0
    max_results = 50
    while True:
        url = f"{base}/rest/api/3/search/jql?jql={quote_plus(jql)}&fields=project,worklog,summary&startAt={start_at}&maxResults={max_results}"
        print("GET", url)
        resp = requests.get(url, headers=headers, timeout=30)
        if resp.status_code // 100 != 2:
            raise RuntimeError(f"Jira API error: {resp.status_code} - {resp.text}")
        root = resp.json()
        issues = root.get('issues', [])
        if not issues:
            break
        all_issues.extend(issues)
        total = int(root.get('total', len(all_issues)))
        returned = int(root.get('maxResults', max_results))
        start_at += returned
        if start_at >= total:
            break
    return all_issues


def main():
    print("Jira Worklog CLI - generates issue-wise worklog CSV")
    base = input(f"Jira base URL [{DEFAULT_BASE}]: ").strip() or DEFAULT_BASE
    email = get_env_or_prompt('EMAIL', 'Email: ')
    api_token = get_env_or_prompt('API_KEY', 'API token: ', hide=True)
    period = input("Period (this, last, all) [this]: ").strip() or 'this'
    if period not in ('this', 'last', 'all'):
        print("Unknown period. Use this, last or all.")
        sys.exit(1)
    out = input("Output file [worklog-report.xlsx]: ").strip() or 'worklog-report.xlsx'
    csv_mode = out.lower().endswith('.csv')

    if period == 'this':
        jql = "worklogAuthor=currentUser() AND worklogDate>=startOfMonth() AND worklogDate<=endOfMonth()"
    elif period == 'last':
        jql = "worklogAuthor=currentUser() AND worklogDate>=startOfMonth(-1) AND worklogDate<=endOfMonth(-1)"
    else:
        jql = "worklogAuthor=currentUser()"

    print("Fetching issues...")
    # Allow running locally without calling Jira by setting MOCK_JIRA=1
    if os.environ.get('MOCK_JIRA') == '1':
        now_iso = datetime.now(pytz.UTC).isoformat()
        user_email = os.environ.get('EMAIL', email)
        issues = [
            {
                'key': 'PROJ-1',
                'fields': {
                    'summary': 'Fix login bug',
                    'project': {'name': 'Project A'},
                    'worklog': {'worklogs': [
                        {'author': {'emailAddress': user_email}, 'timeSpentSeconds': 3600, 'started': now_iso}
                    ]}
                }
            },
            {
                'key': 'PROJ-2',
                'fields': {
                    'summary': 'Add reporting',
                    'project': {'name': 'Project B'},
                    'worklog': {'worklogs': [
                        {'author': {'emailAddress': user_email}, 'timeSpentSeconds': 7200, 'started': now_iso}
                    ]}
                }
            }
        ]
    else:
        try:
            issues = search_issues(base, email, api_token, jql)
        except Exception as ex:
            print("Failed to call Jira:", ex)
            sys.exit(2)

    rng = compute_range(period, pytz.timezone(os.environ.get('TZ', 'UTC')))

    def period_label(period_str: str) -> str:
        tz = pytz.timezone(os.environ.get('TZ', 'UTC'))
        now = datetime.now(tz).date()
        if period_str == 'this':
            return now.strftime('%B %Y')
        if period_str == 'last':
            last = (now.replace(day=1) - timedelta(days=1))
            return last.strftime('%B %Y')
        return period_str

    friendly_period = period_label(period)

    totals: Dict[str, int] = {}
    summaries: Dict[str, str] = {}
    projects: Dict[str, str] = {}
    # collect per-worklog detail rows: tuple(issue_key, summary, project, started_iso, created_iso, seconds, comment)
    details: List[Tuple[str, str, str, str, str, int, str]] = []
    # collect per-day, per-issue aggregates in IST: key = (date_iso, issue_key) -> seconds
    day_totals: Dict[Tuple[str, str], int] = {}
    IST = pytz.FixedOffset(330)  # IST is UTC+5:30

    for issue in issues:
        key = issue.get('key', 'UNKNOWN')
        summary = issue.get('fields', {}).get('summary', '')
        project = issue.get('fields', {}).get('project', {}).get('name', 'UNKNOWN')
        summaries.setdefault(key, summary)
        projects.setdefault(key, project)
        worklogs = issue.get('fields', {}).get('worklog', {}).get('worklogs', [])
        if not isinstance(worklogs, list):
            continue
        issue_seconds = totals.get(key, 0)
        for wl in worklogs:
            author = wl.get('author', {}).get('emailAddress') or wl.get('author', {}).get('name')
            env_email = os.environ.get('EMAIL')
            if env_email and author and author.lower() != env_email.lower():
                continue
            seconds = int(wl.get('timeSpentSeconds') or 0)
            started = wl.get('started')
            if not started:
                continue
            try:
                odt = parser.isoparse(started)
                if odt.tzinfo is None:
                    odt = odt.replace(tzinfo=pytz.UTC)
                if rng is not None:
                    start, end = rng
                    if odt < start or odt > end:
                        continue
                issue_seconds += seconds
                # capture a detail row; prefer 'comment' or 'comment' field if present
                comment = ''
                if 'comment' in wl and wl.get('comment'):
                    c = wl.get('comment')
                    comment = _flatten_text(c)

                # started is odt; convert to IST and store ISO
                started_ist = odt.astimezone(IST)
                started_iso = started_ist.isoformat()

                # try to extract a created timestamp from the worklog entry if present
                created_iso = ''
                if 'created' in wl and wl.get('created'):
                    try:
                        codt = parser.isoparse(wl.get('created'))
                        if codt.tzinfo is None:
                            codt = codt.replace(tzinfo=pytz.UTC)
                        created_iso = codt.astimezone(IST).isoformat()
                    except Exception:
                        created_iso = str(wl.get('created'))

                details.append((key, summary, project, started_iso, created_iso, seconds, comment))
                # aggregate into day_totals using the IST date
                try:
                    date_str = started_ist.date().isoformat()
                    day_totals[(date_str, key)] = day_totals.get((date_str, key), 0) + seconds
                except Exception:
                    # defensive: if date extraction fails, skip day aggregation for that entry
                    pass
            except Exception:
                continue
        totals[key] = issue_seconds

    list_totals = [IssueTotal(k, summaries.get(k, ''), projects.get(k, ''), v) for k, v in totals.items()]
    if period != 'all':
        list_totals = [it for it in list_totals if it.total_seconds > 0]
    list_totals.sort(key=lambda it: it.hours(), reverse=True)

    grand = sum(it.total_seconds for it in list_totals)

    if csv_mode:
        with open(out, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(['Issue Key', 'Summary', 'Project', 'Total Hours', 'Total Days (8h)'])
            for it in list_totals:
                w.writerow([it.key, it.summary, it.project, f"{it.hours():.2f}", f"{it.days():.2f}"])
            w.writerow(['GRAND TOTAL', '', '', f"{grand/3600.0:.2f}", f"{(grand/3600.0)/8.0:.2f}"])
        print(f"Wrote CSV report to {out}")
    else:
        # produce an Excel .xlsx workbook using openpyxl
        try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Alignment
        except Exception:
            print("openpyxl is required to write Excel files. Install with 'pip install openpyxl'.")
            sys.exit(3)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Worklog'

        # Header rows
        ws.append(['Jira Worklog Report'])
        ws.append([f'Period: {friendly_period}', f'Generated: {datetime.now().isoformat()}'])
        ws.append([])
        headers = ['Issue Key', 'Summary', 'Project', 'Total Hours', 'Total Days (8h)']
        ws.append(headers)

        for it in list_totals:
            ws.append([it.key, it.summary, it.project, float(f"{it.hours():.2f}"), float(f"{it.days():.2f}")])

        # Grand total row
        ws.append(['', '', '', '', ''])
        ws.append(['GRAND TOTAL', '', '', float(f"{grand/3600.0:.2f}"), float(f"{(grand/3600.0)/8.0:.2f}")])

        # simple column width adjustments for summary sheet
        col_widths = {}
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for i, cell in enumerate(row, start=1):
                val = cell.value
                if val is None:
                    length = 0
                else:
                    length = len(str(val))
                col_widths[i] = max(col_widths.get(i, 0), length)

        for i, width in col_widths.items():
            ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 8), 60)

        # number formatting for hours/days columns (summary)
        for row in ws.iter_rows(min_row=5, min_col=4, max_col=5, max_row=ws.max_row):
            for cell in row:
                cell.number_format = '0.00'

        # add a Details sheet with per-worklog rows
        details_ws = wb.create_sheet('Details')
        details_headers = ['Issue Key', 'Summary', 'Project', 'Started (IST)', 'Created (IST)', 'Hours', 'Comment']
        details_ws.append(details_headers)
        for d in details:
            ikey, summ, proj, started_iso, created_iso, seconds, comment = d
            hours = round(seconds / 3600.0, 2)
            details_ws.append([ikey, summ, proj, started_iso, created_iso, hours, comment])

        # simple column width adjustments for details sheet
        dcol_widths = {}
        for row in details_ws.iter_rows(min_row=1, max_row=details_ws.max_row):
            for i, cell in enumerate(row, start=1):
                val = cell.value
                if val is None:
                    length = 0
                else:
                    length = len(str(val))
                dcol_widths[i] = max(dcol_widths.get(i, 0), length)

        for i, width in dcol_widths.items():
            details_ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 8), 80)

        # number formatting for hours column in details (Hours is column 6)
        for row in details_ws.iter_rows(min_row=2, min_col=6, max_col=6, max_row=details_ws.max_row):
            for cell in row:
                cell.number_format = '0.00'

        # add a By Day sheet with date-wise per-ticket hours (IST)
        byday_ws = wb.create_sheet('By Day')
        byday_headers = ['Date (IST)', 'Issue Key', 'Project', 'Hours', 'Day Total']
        byday_ws.append(byday_headers)
        # sort by date then issue key for stable ordering
        # prepare a per-date grand total (seconds)
        per_date_seconds: Dict[str, int] = {}
        for (date_str, ikey), secs in day_totals.items():
            per_date_seconds[date_str] = per_date_seconds.get(date_str, 0) + secs

        for (date_str, ikey), secs in sorted(day_totals.items(), key=lambda kv: (kv[0][0], kv[0][1])):
            proj = projects.get(ikey, '')
            hours = round(secs / 3600.0, 2)
            # day_total will be filled into merged cells later; for now put the numeric value on the first row of the date
            byday_ws.append([date_str, ikey, proj, hours, per_date_seconds.get(date_str, 0) / 3600.0])

        # column width adjustments for By Day sheet
        bcol_widths = {}
        for row in byday_ws.iter_rows(min_row=1, max_row=byday_ws.max_row):
            for i, cell in enumerate(row, start=1):
                val = cell.value
                if val is None:
                    length = 0
                else:
                    length = len(str(val))
                bcol_widths[i] = max(bcol_widths.get(i, 0), length)

        for i, width in bcol_widths.items():
            byday_ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 8), 40)

        # number formatting for hours column in By Day (column 4) and Day Total (column 5)
        for row in byday_ws.iter_rows(min_row=2, min_col=4, max_col=5, max_row=byday_ws.max_row):
            for cell in row:
                cell.number_format = '0.00'

        # merge Date (IST) cells so each date appears only once (merge consecutive identical dates)
        # and merge Day Total cells (column 5) across the same spans
        if byday_ws.max_row >= 2:
            merge_start = 2
            prev_date = byday_ws.cell(row=2, column=1).value
            for r in range(3, byday_ws.max_row + 1):
                cur_date = byday_ws.cell(row=r, column=1).value
                if cur_date == prev_date:
                    # continue the current span
                    continue
                # date changed: if span length > 1, merge
                span_end = r - 1
                if merge_start < span_end:
                    # merge Date column
                    byday_ws.merge_cells(start_row=merge_start, start_column=1, end_row=span_end, end_column=1)
                    # merge Day Total column
                    byday_ws.merge_cells(start_row=merge_start, start_column=5, end_row=span_end, end_column=5)
                    # center vertically the merged cells for nicer layout
                    byday_ws.cell(row=merge_start, column=1).alignment = Alignment(vertical='center')
                    byday_ws.cell(row=merge_start, column=5).alignment = Alignment(vertical='center')
                else:
                    # single cell - set vertical center alignment as well
                    byday_ws.cell(row=merge_start, column=1).alignment = Alignment(vertical='center')
                    byday_ws.cell(row=merge_start, column=5).alignment = Alignment(vertical='center')
                # start new span
                merge_start = r
                prev_date = cur_date
            # finalize last span
            if merge_start < byday_ws.max_row:
                byday_ws.merge_cells(start_row=merge_start, start_column=1, end_row=byday_ws.max_row, end_column=1)
                byday_ws.merge_cells(start_row=merge_start, start_column=5, end_row=byday_ws.max_row, end_column=5)
                byday_ws.cell(row=merge_start, column=1).alignment = Alignment(vertical='center')
                byday_ws.cell(row=merge_start, column=5).alignment = Alignment(vertical='center')
            else:
                byday_ws.cell(row=merge_start, column=1).alignment = Alignment(vertical='center')
                byday_ws.cell(row=merge_start, column=5).alignment = Alignment(vertical='center')

        wb.save(out)
        print(f"Wrote Excel report to {out} (sheets: {', '.join(wb.sheetnames)})")


if __name__ == '__main__':
    main()
